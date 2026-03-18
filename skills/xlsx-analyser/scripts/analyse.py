#!/usr/bin/env python3
"""
xlsx-analyser/scripts/analyse.py

Reads an .xlsx, .xls, or .csv file and outputs a structured JSON analysis
report to stdout. Requires openpyxl (for xlsx/xls) or uses stdlib csv module
as a fallback for .csv files.

Usage:
    python analyse.py <file_path> [--sheet <sheet_name>]

Output (stdout): JSON with the following structure:
    {
      "file": "path/to/file.xlsx",
      "format": "xlsx" | "csv",
      "sheets": [
        {
          "name": "Sheet1",
          "total_rows": 100,
          "total_columns": 5,
          "dimensions": "100 rows x 5 columns",
          "columns": [
            {
              "name": "col_name",
              "index": 0,
              "inferred_type": "integer" | "float" | "boolean" | "date" | "text" | "mixed" | "empty",
              "total": 100,
              "non_empty": 98,
              "null_count": 2,
              "null_pct": 2.0,
              "unique_count": 15,
              "unique_values": ["a", "b"] | null,   # only when unique_count <= 5
              "stats": { ... },                       # type-specific stats
              "flags": ["high_nulls", "constant", "unnamed"]
            }
          ],
          "issues": ["..."]
        }
      ],
      "error": null | "error message"
    }
"""

from __future__ import annotations
import sys
import json
import csv
import re
from pathlib import Path
from collections import Counter
from datetime import datetime, date


# ── type inference ───────────────────────────────────────────────────────────

_BOOL_TRUE  = {"true", "yes", "1", "y"}
_BOOL_FALSE = {"false", "no", "0", "n"}
_BOOL_ALL   = _BOOL_TRUE | _BOOL_FALSE
_DATE_PATTERNS = [
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),          # YYYY-MM-DD
    re.compile(r"^\d{2}/\d{2}/\d{4}$"),           # DD/MM/YYYY or MM/DD/YYYY
    re.compile(r"^\d{4}/\d{2}/\d{2}$"),           # YYYY/MM/DD
]


def _infer_type(values: list) -> str:
    """Infer the dominant type of a list of non-null raw values."""
    if not values:
        return "empty"

    str_vals = [str(v).strip() for v in values]

    # boolean check
    if all(s.lower() in _BOOL_ALL for s in str_vals):
        return "boolean"

    # date check (native date/datetime objects from openpyxl, or string patterns)
    native_dates = sum(1 for v in values if isinstance(v, (date, datetime)))
    if native_dates == len(values):
        return "date"
    if all(any(p.match(s) for p in _DATE_PATTERNS) for s in str_vals):
        return "date"

    # numeric checks
    int_count = float_count = 0
    for v in values:
        if isinstance(v, bool):
            continue
        if isinstance(v, int):
            int_count += 1
        elif isinstance(v, float):
            float_count += 1
        else:
            s = str(v).strip().replace(",", "")
            try:
                parsed = float(s)
                if "." in s or "e" in s.lower():
                    float_count += 1
                else:
                    int_count += 1
            except ValueError:
                return "text"

    total = len(values)
    if int_count + float_count == total:
        return "float" if float_count > 0 else "integer"

    return "text"


def _to_number(v) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", ""))
    except (ValueError, TypeError):
        return None


# ── per-column statistics ────────────────────────────────────────────────────

def _safe_round(v: float, n: int = 2) -> float:
    try:
        return round(v, n)
    except Exception:
        return v


def _compute_stats(col_type: str, non_null_values: list) -> dict:
    stats: dict = {}
    if not non_null_values:
        return stats

    if col_type in ("integer", "float"):
        nums = [n for n in (_to_number(v) for v in non_null_values) if n is not None]
        if nums:
            stats["min"]    = _safe_round(min(nums))
            stats["max"]    = _safe_round(max(nums))
            stats["mean"]   = _safe_round(sum(nums) / len(nums))
            sorted_nums = sorted(nums)
            n = len(sorted_nums)
            mid = n // 2
            stats["median"] = _safe_round(
                sorted_nums[mid] if n % 2 else (sorted_nums[mid - 1] + sorted_nums[mid]) / 2
            )
            if len(nums) >= 2:
                mean_v = stats["mean"]
                variance = sum((x - mean_v) ** 2 for x in nums) / (len(nums) - 1)
                stats["std_dev"] = _safe_round(variance ** 0.5)

    elif col_type == "boolean":
        str_vals = [str(v).strip().lower() for v in non_null_values]
        stats["true_count"]  = sum(1 for s in str_vals if s in _BOOL_TRUE)
        stats["false_count"] = sum(1 for s in str_vals if s in _BOOL_FALSE)

    elif col_type == "text":
        lengths = [len(str(v)) for v in non_null_values]
        stats["avg_length"] = _safe_round(sum(lengths) / len(lengths))
        stats["min_length"] = min(lengths)
        stats["max_length"] = max(lengths)
        # top 3 most common values
        counter = Counter(str(v) for v in non_null_values)
        stats["top_values"] = [{"value": v, "count": c} for v, c in counter.most_common(3)]

    elif col_type == "date":
        date_strs = [str(v) for v in non_null_values]
        stats["earliest"] = min(date_strs)
        stats["latest"]   = max(date_strs)

    return stats


# ── column analysis ──────────────────────────────────────────────────────────

def _analyse_column(name: str, index: int, values: list) -> dict:
    total      = len(values)
    non_null   = [v for v in values if v is not None and str(v).strip() != ""]
    null_count = total - len(non_null)
    null_pct   = _safe_round(null_count / total * 100) if total else 0.0

    col_type   = _infer_type(non_null)
    unique_raw = list({str(v) for v in non_null})
    unique_count = len(unique_raw)

    flags: list[str] = []
    if null_pct > 20:
        flags.append("high_nulls")
    if unique_count == 1 and non_null:
        flags.append("constant_column")
    if not name or name.lower().startswith("unnamed") or name.lower().startswith("column_"):
        flags.append("unnamed_column")

    return {
        "name":          name,
        "index":         index,
        "inferred_type": col_type,
        "total":         total,
        "non_empty":     len(non_null),
        "null_count":    null_count,
        "null_pct":      null_pct,
        "unique_count":  unique_count,
        "unique_values": sorted(unique_raw)[:10] if unique_count <= 5 else None,
        "stats":         _compute_stats(col_type, non_null),
        "flags":         flags,
    }


# ── sheet analysis ───────────────────────────────────────────────────────────

def _analyse_rows(headers: list[str], data_rows: list[list]) -> list[dict]:
    columns = []
    for idx, header in enumerate(headers):
        col_values = [
            row[idx] if idx < len(row) else None
            for row in data_rows
        ]
        columns.append(_analyse_column(header, idx, col_values))
    return columns


def _sheet_issues(columns: list[dict]) -> list[str]:
    issues: list[str] = []
    for col in columns:
        for flag in col["flags"]:
            if flag == "high_nulls":
                issues.append(f"Column '{col['name']}' has {col['null_pct']}% null values.")
            elif flag == "constant_column":
                uv = col["unique_values"]
                val = uv[0] if uv else "?"
                issues.append(f"Column '{col['name']}' is constant (only value: '{val}').")
            elif flag == "unnamed_column":
                issues.append(f"Column at index {col['index']} has no meaningful name.")
    return issues


# ── xlsx reader ──────────────────────────────────────────────────────────────

def _analyse_xlsx(path: Path, target_sheet: str | None = None) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {
            "file": str(path),
            "error": (
                "openpyxl is not installed. "
                "Install it with: pip install openpyxl"
            ),
        }

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception as e:
        return {"file": str(path), "error": f"Failed to open workbook: {e}"}

    sheet_names = wb.sheetnames
    if target_sheet and target_sheet not in sheet_names:
        return {
            "file":  str(path),
            "error": f"Sheet '{target_sheet}' not found. Available: {sheet_names}",
        }

    sheets_to_analyse = [target_sheet] if target_sheet else sheet_names
    sheets_result = []

    for sname in sheets_to_analyse:
        ws = wb[sname]
        try:
            all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        except Exception as e:
            sheets_result.append({"name": sname, "error": str(e)})
            continue

        if not all_rows:
            sheets_result.append({"name": sname, "empty": True, "total_rows": 0, "total_columns": 0})
            continue

        # strip fully-empty trailing rows
        while all_rows and all(v is None for v in all_rows[-1]):
            all_rows.pop()

        headers   = [str(h) if h is not None else f"Column_{i+1}" for i, h in enumerate(all_rows[0])]
        data_rows = all_rows[1:]
        columns   = _analyse_rows(headers, data_rows)

        sheets_result.append({
            "name":           sname,
            "dimensions":     f"{len(data_rows)} rows x {len(headers)} columns",
            "total_rows":     len(data_rows),
            "total_columns":  len(headers),
            "columns":        columns,
            "issues":         _sheet_issues(columns),
        })

    wb.close()
    return {
        "file":        str(path),
        "format":      path.suffix.lstrip(".").lower(),
        "sheet_names": sheet_names,
        "sheets":      sheets_result,
        "error":       None,
    }


# ── csv reader ───────────────────────────────────────────────────────────────

def _analyse_csv(path: Path) -> dict:
    try:
        with path.open(newline="", encoding="utf-8-sig") as f:
            sample = f.read(4096)
            f.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            reader  = csv.reader(f, dialect)
            all_rows = [row for row in reader]
    except Exception:
        try:
            with path.open(newline="", encoding="latin-1") as f:
                reader   = csv.reader(f)
                all_rows = list(reader)
        except Exception as e:
            return {"file": str(path), "error": f"Failed to read CSV: {e}"}

    if not all_rows:
        return {"file": str(path), "format": "csv", "sheets": [{"name": "Sheet1", "empty": True}]}

    while all_rows and all(v.strip() == "" for v in all_rows[-1]):
        all_rows.pop()

    headers   = [h.strip() or f"Column_{i+1}" for i, h in enumerate(all_rows[0])]
    data_rows = all_rows[1:]
    columns   = _analyse_rows(headers, data_rows)

    sheet = {
        "name":          "Sheet1",
        "dimensions":    f"{len(data_rows)} rows x {len(headers)} columns",
        "total_rows":    len(data_rows),
        "total_columns": len(headers),
        "columns":       columns,
        "issues":        _sheet_issues(columns),
    }

    return {
        "file":        str(path),
        "format":      "csv",
        "sheet_names": ["Sheet1"],
        "sheets":      [sheet],
        "error":       None,
    }


# ── entry point ──────────────────────────────────────────────────────────────

def main() -> None:
    args = sys.argv[1:]
    if not args:
        print(json.dumps({"error": "Usage: python analyse.py <file_path> [--sheet <name>]"}))
        sys.exit(1)

    file_path = Path(args[0])
    target_sheet: str | None = None
    if "--sheet" in args:
        idx = args.index("--sheet")
        if idx + 1 < len(args):
            target_sheet = args[idx + 1]

    if not file_path.exists():
        print(json.dumps({"error": f"File not found: {file_path}"}))
        sys.exit(1)

    ext = file_path.suffix.lower()
    if ext == ".csv":
        result = _analyse_csv(file_path)
    elif ext in (".xlsx", ".xls", ".xlsm", ".xlsb"):
        result = _analyse_xlsx(file_path, target_sheet)
    else:
        print(json.dumps({"error": f"Unsupported file type: {ext}. Supported: .xlsx, .xls, .xlsm, .csv"}))
        sys.exit(1)

    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()
